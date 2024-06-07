import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, timedelta

file_path = 'C:/Users/prr8ca/Desktop/MailQT/BLOCO K - Planejamento.xlsx'
df = pd.read_excel(file_path, sheet_name='FUP Prazos', engine='openpyxl')

wb = load_workbook(file_path)
ws = wb['FUP Prazos']

def buscar_fornecedores(): #Retorna uma lista com todos os fornecedores

    razao_social_unicos = df['Razão Social'].unique()
    razao_social_lista = razao_social_unicos.tolist()
    return razao_social_lista

def email_do_fornecedor(nome_fornecedor): #Retorna uma Lista com os email do fornecedor
    filtro_fornecedor = df['Razão Social'] == nome_fornecedor
    email = df.loc[filtro_fornecedor, df.columns[26]].iloc[0] if filtro_fornecedor.any() else None
    return email


def ver_prazos_vencidos(nome_fornecedor):
    prazos_vencidos = []
    filtro_fornecedor = df['Razão Social'] == nome_fornecedor

    for index, _ in df[filtro_fornecedor].iterrows():
        excel_linha_index = index + 2  
        prazo_cell = ws[f'N{excel_linha_index}']  
        cell_content = prazo_cell.value

        if cell_content is None or cell_content + timedelta(days=1) < datetime.now():
            prazos_vencidos.append(excel_linha_index)

    print(prazos_vencidos)
    return(prazos_vencidos)


def ver_informacoes_necessarias(numeros_linha):
    informacoes = []
    for numero_linha in numeros_linha:
        linha = df.iloc[numero_linha - 2]  
        informacao_linha = [valor.strftime("%d/%m/%Y") if isinstance(valor, pd.Timestamp) else valor for i, valor in enumerate(linha) if i in [0, 1, 2, 3, 4, 5, 6, 7, 8, 13]] 
        informacoes.append(informacao_linha)
    
    print(informacoes)
    return informacoes



    

